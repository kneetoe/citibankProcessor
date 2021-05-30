import openpyxl
from openpyxl.styles.builtins import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
import math

def char_range(c1, c2):
    """Generates the characters from `c1` to `c2`, inclusive."""
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)

def autoSize(worksheet):
    i = 0
    for column_cells in worksheet.columns:
        if(i==3 or i == 4):
            multp = 2
        else:
            multp = 1.25
        length = max(len(str(cell.value)) for cell in column_cells) * multp
        worksheet.column_dimensions[column_cells[0].column_letter].width = length
        i = i + 1


def sums(data):
    sums = []
    sum1 = 0
    sum2 = 0
    sum3 = 0
    sum4 = 0
    i=1
    length = len(data[0])
    while(i < length):
        if(data[5][i] == 'CARL DOE' and data[3][i] != '0'):
            sum1 = sum1 + float(data[3][i])
        elif(data[5][i] == 'JANET DOE' and data[3][i] != '0'):
            sum2 = sum2 + float(data[3][i])
        elif(data[5][i] == 'JOHN DOE' and data[3][i] != '0'):
            sum3 = sum3 + float(data[3][i])
        elif(str(data[3][i]) != '0'):
            sum4 = (math.fsum([sum4, float(data[3][i])]))
        i = i + 1
        
    lst = []
    lst.append('CARL DOE')
    lst.append('{:.2f}'.format(sum1))
    sums.append(lst)
    
    lst = []
    lst.append('JANET DOE')
    lst.append('{:.2f}'.format(sum2))
    sums.append(lst)
    
    lst = []
    lst.append('JOHN DOE')
    lst.append('{:.2f}'.format(sum3))
    sums.append(lst)
    
    lst = []
    lst.append('JESSICA DOE')
    lst.append('{:.2f}'.format(sum4))
    sums.append(lst)

    return sums

def findName(name):
    year = findYear(name)
    finalName = 'Unnamed sheet ' + year
    if('jan' in name.lower()):
        finalName = 'January ' + year
    elif('feb' in name.lower()):
        finalName = 'Febuary ' + year
    elif('mar' in name.lower()):
        finalName = 'March ' + year
    elif('apr' in name.lower()):
        finalName = 'April ' + year
    elif('may' in name.lower()):
        finalName = 'May ' + year
    elif('jun' in name.lower()):
        finalName = 'June ' + year
    elif('jul' in name.lower()):
        finalName = 'July ' + year
    elif('aug' in name.lower()):
        finalName = 'August ' + year
    elif('sep' in name.lower()):
        finalName = 'September ' + year
    elif('oct' in name.lower()):
        finalName = 'October ' + year
    elif('nov' in name.lower()):
        finalName = 'Novermber ' + year
    elif('dec' in name.lower()):
        finalName = 'December ' + year
    return finalName

def findYear(name):
    finalYear = '?'
    if('2020' in name):
        finalYear = '2020'
    elif('2021' in name):
        finalYear = '2021'
    elif('2022' in name):
        finalYear = '2022'
    elif('2023' in name):
        finalYear = '2023'
    elif('2024' in name):
        finalYear = '2024'
    elif('2025' in name):
        finalYear = '2025'
    else:
        finalYear = 'Time to update program'
    return finalYear
    

#populate new workbook
def populate(arr, path, name):
    book = openpyxl.load_workbook(path + '\currentYear.xlsx')

    ws_name = findName(name)
    j = 0
    while(j < len(book.sheetnames)):
        if(ws_name in book.sheetnames[j] and not 'copy' in book.sheetnames[j]):
            ws_name = ws_name + " copy"
        j = j + 1
    
    ws1 = book.create_sheet(ws_name)

    i = 0
    arr[6][0] = 'Category'
    arr[8][0] = 'Comments'

    person = 'none'
    lister = []

    length = len(arr[0])
    while(i < length):
        ws1[('A'+ str(i+1))] = arr[0][i]
        ws1[('B'+ str(i+1))] = arr[1][i]
        ws1[('C'+ str(i+1))] = arr[2][i]
        if(arr[3][i] == '0'):
            ws1[('D'+ str(i+1))] = ''
        elif(not i == 0):
            ws1[('D'+ str(i+1))] = float(arr[3][i])
            ws1[('D'+ str(i+1))].style = 'Currency'
        else:
            ws1[('D'+ str(i+1))] = (arr[3][i])
            
        if(arr[4][i] == '0'):
            ws1[('E'+ str(i+1))] = ''
        elif(not i == 0):
            ws1[('E'+ str(i+1))] = float(arr[4][i])
            ws1[('E'+ str(i+1))].style = 'Currency'
            
        else:
            ws1[('E'+ str(i+1))] = arr[4][i]
            
        ws1[('F'+ str(i+1))] = arr[5][i]
        '''if(arr[5][i] != person and i > 0):
            if(i!=1):
              lister.append(cost) 
            cost = []
            cost.append(arr[5][i])
            cost.append(float(arr[3][i]))
        elif(i > 0):
            cost[1] = cost[1] + arr[3][i]'''
        
        if(arr[6][i] == 'NONE'):
            arr[6][i] = ''
        ws1[('G'+ str(i+1))] = arr[6][i]

        if(arr[8][i] == 'NONE'):
            arr[8][i] = ''
        ws1[('H'+ str(i+1))] = arr[8][i]

        
        #ws1[('H'+ str(i+1))] = arr[7][i]
        i = i + 1
    sum_lst = sums(arr)
    i = i + 2

    ws1[('A'+ str(i))] = 'TOTALS'
    i = i + 1

    j = 0
    while(j < len(sum_lst)):
        ws1[('A'+ str(i))] = sum_lst[j][0]
        ws1[('B'+ str(i))] = float(sum_lst[j][1])
        ws1[('B'+ str(i))].style = 'Currency'
        j = j + 1
        i = i + 1
        
    

    
    autoSize(ws1)
    tab = Table(displayName="Table1", ref="A1:G" + str(length))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    
        
    book.save(path + '\currentYear.xlsx') 
